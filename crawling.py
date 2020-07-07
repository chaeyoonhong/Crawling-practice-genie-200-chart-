import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
load_wb = load_workbook("chart.xlsx",data_only=True)
load_ws = load_wb["Sheet1"]
# headers 사이트 접근 제한 해제
headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.116 Safari/537.36'}
page=["1","2","3","4"]
#tr class 가져옴
for j in range(4):
    data = requests.get("https://www.genie.co.kr/chart/top200?ditc=D&ymd=20200706&hh=14&rtm=N&pg="+page[j],headers=headers)
    soup = BeautifulSoup(data.text,'html.parser')
    trs = soup.select ('#body-content > div.newest-list > div > table > tbody > tr')
    title=[]
    singer=[]
    rank=[]
    i=0
#title, singer 에 정보 입력
    for tr in trs:
        title.append(tr.select_one('td.info>a.title.ellipsis').text.strip())
        singer.append(tr.select_one('td.info>a.artist.ellipsis').text.strip())
        rank.append(tr.select_one('td.number').text.strip())
        load_ws.cell(50*j+i+1,1,rank[i])
        load_ws.cell(50*j+i+1,2,title[i])
        load_ws.cell(50*j+i+1,3,singer[i])
        i=i+1
load_wb.save("chart.xlsx")
